import openpyxl

def get_data(filename, sheet_name):
    data=[]
    workbook=openpyxl.load_workbook("./DataFiles/"+filename)
    sheet=workbook[sheet_name]

    for r in range(2, sheet.max_row+1):
        row_list = []
        for c in range(1, sheet.max_column+1):
            row_list.append(sheet.cell(r, c).value)
        data.append(row_list)

    return data